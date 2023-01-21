#! python3

#import pandas as pd
import os, sys
import json
import openpyxl # for .xlsx files
import xlrd # for .xls files
#from datetime import datetime as dt
import traceback
import keyboard

class TA_Payroll_Upload():
    def __init__(self, param_file=str):
        """
        Takes values from three sheets and adds them to an output file
        Used for uploading conrtibution and YTD values from Quickbooks to Transamerica 401k
        """
        title = None # Used in debugging
        try:
            # Set variables
            self.scriptpath = os.path.dirname(os.path.realpath(sys.argv[0])) + '\\'
            #tstamp = dt.now().strftime("%Y-%m-%d_%H-%M-%S")
            self.vars = self.load_json(param_file)

            # Fill in data in self.vars
            for key in self.vars:
                try:
                    # Replace 'scriptpath' str in json with actual script path
                    if self.vars[key]['path'] == 'scriptpath':
                        self.vars[key]['path'] = self.scriptpath
                    
                    # Add full file path of latest file matching key to ['file']
                    self.vars[key]['file'] = self.return_paths(key)
                    
                    # Add data to files that are "read_files"
                    if self.vars[key]['read_file']:
                        # Add xls data to ['data'], uses xlrd for legacy excel files (.xls) 
                        if self.vars[key]['type'] =='.xls':
                            self.vars[key]['data'] = self.get_xls_data(key)
                        else:
                            sys.exit(input('\nError : Unable to read non .xls files at this time. Tell Kyle to add other file types if you need them.'))
                except KeyError:
                    continue

            # Add the data from "read_files" to the output file
            self.parse_data()

            # Open the final file
            print('Opening file "' + self.vars['payroll_file']['file'] + '"')
            os.system(str('"' + self.vars['payroll_file']['file'] + '"'))
            
        except PermissionError:
            sys.exit(input('\n\nError : Cannot access the output file, it is likely open. Please close the file and try again.'))
        
        except Exception as e:
            print('\n\n+++ An error has occured - Get Kyle! +++\n\n' + traceback.format_exc())
            print('--- Press "v" for debug variable values, or any other key to quit ---')
            while True:
                if keyboard.read_key() == "v":
                    self.vars['debug']['active'] = True
                    title = ' '
                    break
                else:
                    sys.exit()

        # Optional debugging. Called if varibles.json ['debug']['active'] = true
        if self.vars['debug']['active']:
            localvars = None
            if self.vars['debug']['locals']:
                localvars = locals()
            scope = None
            if self.vars['debug']['scope']:
                scope = dir()
            self._debug(title=title, close=False, localvars=localvars, scope=scope, globalvars=self.vars['debug']['globals'], **self.vars['debug']['kwargs'])
        
        sys.exit(input('\nPRESS ANY KEY TO CLOSE'))
        
    def load_json(self, filename):
        """Return data from JSON file as a dict"""
        try:
            with open(self.scriptpath + filename) as f:
                data = json.load(f)
        except Exception as e:
            sys.exit('Unable to load JSON data from "' + filename + '". ' + str(e))
        return data

    def return_paths(self, searchkeyword):
        """Find the file to be edited"""
        homepath = 'C:' + os.environ["HOMEPATH"]
        searchfile = self.vars[searchkeyword]["namekeyword"]
        searchtype = self.vars[searchkeyword]["type"]
        searchpath = self.vars[searchkeyword]["path"]
        if not searchpath.startswith('C:'):
            searchpath = homepath + '\\'+ searchpath +'\\'
        return self._file_latest_file(searchfile, searchtype, searchpath)

    def _file_latest_file(self, searchfile=str, searchtype=str, searchpath=str):
        """Return the most recent file matching search criteria"""
        allfiles = os.listdir(searchpath)
        matchfiles = [os.path.join(searchpath, basename) for basename in allfiles if basename.endswith(searchtype) and searchfile in basename]
        try:
            foundfile = max(matchfiles, key=os.path.getctime)
        except ValueError:
            print('Could not find file containing "' + searchfile + '" in folder "' + searchpath + '". Ensure that you have downloaded the correct file, or adjust variable.json search parameters.')
            p = sys.exit(input())
        return foundfile

    def get_xls_data(self, file):
        """
        Return a list of vaules in all rows from columns definied by self.vars.[file][cols]=list
        Data returned will contain ', ' in column 0, and no digits, which ensures the first list [0] is a user name
        This is a hard-coded solution, so this will have to be reworked if major changes in the QB export process occurs
        """
        print('Reading file "' + self.vars[file]['file'] + '"... ', end='', flush=True)
        wb = xlrd.open_workbook(self.vars[file]['file'])
        ws = wb.sheet_by_index(self.vars[file]['sheet'])
        users = []
        # Interate through all rows
        for r in range (0, ws.nrows):    
            user_data = []
            v = ws.cell(r, 0).value
            
            # If the first column has a string that contains ', ' and no numbers, its a name
            if ', ' in v and not any(char.isdigit() for char in v):
                # Remove middle names for data comparison later
                first = v.split(', ')[1].split(' ')[0].strip() # Remove whitespace
                last = v.split(', ')[0].strip()
                full = last + ', ' + first
                user_data.append(full)
                for c in self.vars[file]['data_cols']:
                    val = ws.cell(r, c).value
                    user_data.append(val)
                users.append(user_data)
        print('Complete\n')
        return users

    def parse_data(self):
        """
        Matches names (read_file[data][0]) with the names in the output file
        Adds data from the next indexes to the target columns (read_file[target_cols]) in the output file
        """
        print('Reading file "' + self.vars['payroll_file']['file'] + '"... ', end='', flush=True)
        wb = openpyxl.load_workbook(self.vars['payroll_file']['file'])
        ws = wb.active
        
        # Create a list of all names in payroll_summary to compare with payroll_file to add missing employee data
        print('\n\nSearching for employee discrepancies... ', end='', flush=True)
        self.summary_names = []
        self.missing_names = []
        for name in self.vars['payroll_summary']['data']:
            last = name[0].split(', ')[0].strip() #remove whitespace
            first = name[0].split(', ')[1]
            # Do not include middle initials : they are not consistent across workbooks
            first = first.split(' ')[0].strip()
            full = last.upper() + ', ' + first.upper()
            self.summary_names.append(full)

        # Iterate over names in sheet and remove names from the summary_names list that are present, and add names not found in Payroll Summary to missin_names
        for row in ws.iter_rows():
            fullname = row[4].value.strip() + ', ' + row[2].value.strip()
            if fullname in self.summary_names:
                self.summary_names.remove(fullname)
            else:
                self.missing_names.append(fullname)

        if not self.summary_names:
            print('no new employees found.')
        else:
            print('new employee(s) added:')
            for name in self.summary_names:
                print(name)
        
        # List employees who were on the Payroll File, but not the Payroll Summary
        if self.missing_names:
            self.missing_names.remove('Last Name, First Name')
            if self.missing_names:
                print('\nNo data in Payroll Summary for the following employees:')
                for name in self.missing_names:
                    print(name)        

        # Format new employee names to be added to bottom of sheet
        max_row = len(ws['C'])
        for name in self.summary_names:
            max_row +=1
            # Add company identifier to column A
            ws.cell(row=max_row, column=1).value = 28
            
            # Parse name, reformat, and add to ws at the next open row
            name = name.split()
            last = name[0].split(',')[0]
            ws.cell(row=max_row, column=5).value = last
            first = name[1]
            ws.cell(row=max_row, column=3).value = first

        # Once you have the updated list of employees on the sheet, create new records and add them to the payroll_file sheet
        print('\nUpdating "' + self.vars['payroll_file']['file'] + '"... ', end='', flush=True)
        for row in ws.iter_rows():
            fullname = row[4].value.strip() + ', ' + row[2].value.strip()
            
            # Check if the name each read_file in self.vars matches the name in the sheet row
            for file in self.vars:
                try:
                    if self.vars[file]['read_file']:
                        for employee in self.vars[file]['data']:
                            if employee[0].upper() == fullname:
                                read_index = 1 # list index for first data point is always 1 (0 is employee name)
                                
                                # Add data from index into target column 
                                for col in self.vars[file]['target_cols']:
                                    # TODO Comment - this may be added in the future    
                                    # If the cell already has data, add the new data to it
                                    #add_value = 0
                                    #if row[col].value:
                                    #    try:
                                    #        add_value = float(row[col].value)
                                    #        print(f'existing value ++ {add_value=}')
                                    #    except ValueError:
                                    #        continue
                                    row[col].value = employee[read_index]
                                    read_index += 1
                except KeyError:
                    continue        

        wb.save(self.vars['payroll_file']['file'])
        print('Complete\n')

    def _print_list(self, lst, tab=''):
        """Format a list for easier reading"""
        for i in lst:
            if type(i) is dict:
                self._print_dict(i)
            if type(i) is list:
                print(tab, i)
                self._print_list(i, tab=tab + '     ')
            print(tab, i)  

    def _print_dict(self, d, tab =''):
        """Format a dict for easier reading"""
        for key, value in d.items():
            if key == 'kwargs':
                return True

            if type(value) is dict:
                print('\n', tab, key, ' : ')
                self._print_dict(value, tab=tab + '     ')
                continue
            if type(value) is list:
                if len(value) < 4:
                    print(tab, key, ' : ', value)
                    continue
                else:
                    print(tab, key, ' : ')
                    tab = tab + '     '
                    for v in value:
                        print(tab, v)
                    tab = '     '
                    continue
            print(tab, key,' : ', value)

    def _debug(self, title=None, close=True, localvars=None, scope=None, globalvars=False, **kwargs):
        """
        Print passed variables in terminal
        Useful for .exe scripts 
        """
        if not title:
            title = '\n===== DEBUG MODE =====\nTo turn off debug mode, change "debug" : "active" \nvariable to false in "variables.json" file.'

        print(title)
        
        if globalvars:
            print('\n... GLOBALS ...')
            for key in globals():
                print(key + ' : ' + str(globals()[key]))
        
        if scope:
            print('\n... SCOPE ...')
            for v in scope:
                try:
                    val = eval(v)
                    print(v, " : ", end='', flush=True)
                    print('type=', type(val), ' : ', end='', flush=True)
                    print('value=', val)
                    continue
                except:
                    print(v, ' : type=', type(v))
                    pass
        
        if localvars:
            print('\n... LOCALS ...')
            for key in localvars:
                print(key + ' : ' + str(localvars[key]))
        
        if kwargs:
            # Variable values called in variables.json ['debug']['kwargs']
            print('\n... KWARGS ...')
            for key, value in kwargs.items():
                val = eval(value)
                if type(val) is dict:
                    skip = self._print_dict(val)
                    continue
                if type(val) is list:
                    print('\n', key, ' : ')
                    self._print_list(val, tab='     ')
                else:
                    print('\n', key, '=', eval(value))

        if close:
            sys.exit(input('\n--- PRESS ANY KEY TO CLOSE ---'))

if __name__ == '__main__':
    TA_Payroll_Upload('variables.json')